"""Build the dashboard preview PNG and the Excel workbook from the **real**
CRM analysis (UCI Online Retail II, CC BY 4.0 — see
`data/REAL_DATA_PROVENANCE.md`).

Replaces the previous hand-drawn, simulated-campaign preview: the hero
image now shows the real RFM segmentation, CLV-by-country and cohort
retention computed by `crm_retention.py`. Every text cell written to the
workbook is still neutralised against spreadsheet formula/DDE injection
(kept verbatim, with its test).

matplotlib (Agg, headless) + openpyxl.
"""

from __future__ import annotations

from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt  # noqa: E402
from matplotlib import font_manager  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill  # noqa: E402

import crm_retention  # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
PREVIEW_PATH = ROOT / "assets" / "dashboard_preview.png"
WORKBOOK_PATH = ROOT / "dashboard" / "crm_dashboard.xlsx"

FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")

INK = "#20242E"
MUTE = "#657184"
ACCENTS = ["#1F8A70", "#2F5DA8", "#B5651D", "#6C4AB6", "#C0392B", "#117864", "#7D6608"]


def neutralize_spreadsheet_formula(value: str) -> str:
    if value.startswith(FORMULA_PREFIXES):
        return "'" + value
    return value


def _font(size: int, weight: str = "normal"):
    return font_manager.FontProperties(size=size, weight=weight)


def build_preview(result: dict[str, object]) -> None:
    PREVIEW_PATH.parent.mkdir(parents=True, exist_ok=True)
    rfm = result["rfm"]
    segments = rfm["segments"]
    clv = result["clv_by_country"]["ranked"]
    cohorts = result["cohort_retention"]["cohorts"][:12]
    max_off = result["cohort_retention"]["max_offset"]

    fig = plt.figure(figsize=(16, 10), dpi=100)
    fig.patch.set_facecolor("white")
    gs = fig.add_gridspec(
        3, 2, height_ratios=[0.9, 2.1, 2.1],
        hspace=0.5, wspace=0.22,
        left=0.135, right=0.965, top=0.92, bottom=0.085,
    )

    # ---- header / KPIs ----
    fig.text(0.06, 0.965, "CRM & Retention — Real Data",
             fontproperties=_font(26, "bold"), color=INK)
    fig.text(0.06, 0.94,
             "UCI Online Retail II (CC BY 4.0) · "
             f"reference date {result['reference_date']} · "
             "real UK e-commerce, 2009-2011",
             fontproperties=_font(13), color=MUTE)

    top_seg = max(segments, key=lambda s: s["revenue_eur"])
    kpis = [
        (f"{result['orders']:,}", "real orders"),
        (f"{rfm['total_customers']:,}", "customers"),
        (f"{top_seg['customer_share']:.0%}", f"are {top_seg['segment']}"),
        (f"EUR {top_seg['revenue_eur']/1e6:.1f}M", f"{top_seg['segment']} revenue"),
    ]
    ax_k = fig.add_subplot(gs[0, :])
    ax_k.axis("off")
    for i, (big, small) in enumerate(kpis):
        x = i / len(kpis) + 0.01
        ax_k.text(x, 0.62, big, transform=ax_k.transAxes,
                  fontproperties=_font(28, "bold"), color=ACCENTS[i % len(ACCENTS)])
        ax_k.text(x, 0.18, small, transform=ax_k.transAxes,
                  fontproperties=_font(14), color=MUTE)

    # ---- RFM segments (customers + revenue) ----
    ax1 = fig.add_subplot(gs[1, 0])
    segs = sorted(segments, key=lambda s: s["customers"])
    names = [s["segment"] for s in segs]
    custs = [s["customers"] for s in segs]
    ax1.barh(names, custs, color=ACCENTS[0])
    for i, s in enumerate(segs):
        ax1.text(s["customers"], i, f"  EUR {s['revenue_eur']/1e6:.1f}M",
                 va="center", fontproperties=_font(11), color=MUTE)
    ax1.set_title("RFM lifecycle segments (customers; revenue labelled)",
                  fontproperties=_font(15, "bold"), color=INK, loc="left", pad=10)
    ax1.tick_params(labelsize=11)
    ax1.set_xlabel("customers", fontproperties=_font(11), color=MUTE)
    for sp in ("top", "right"):
        ax1.spines[sp].set_visible(False)

    # ---- CLV by country ----
    ax2 = fig.add_subplot(gs[1, 1])
    cnames = [c["country"] for c in clv]
    cvals = [c["historical_clv_eur"] for c in clv]
    ax2.bar(cnames, cvals, color=ACCENTS[1])
    for i, c in enumerate(clv):
        ax2.text(i, c["historical_clv_eur"], f"EUR {c['historical_clv_eur']:,.0f}",
                 ha="center", va="bottom", fontproperties=_font(10), color=MUTE)
    ax2.set_title("Historical CLV by country (≥ %d customers)"
                  % result["clv_by_country"]["min_customers"],
                  fontproperties=_font(15, "bold"), color=INK, loc="left", pad=10)
    ax2.tick_params(labelsize=11)
    ax2.set_ylabel("EUR / customer", fontproperties=_font(11), color=MUTE)
    for sp in ("top", "right"):
        ax2.spines[sp].set_visible(False)

    # ---- cohort retention heatmap ----
    ax3 = fig.add_subplot(gs[2, :])
    grid = [c["retention"] for c in cohorts]
    im = ax3.imshow(grid, aspect="auto", cmap="YlGn", vmin=0, vmax=1)
    ax3.set_xticks(range(max_off + 1))
    ax3.set_xticklabels([f"M{k}" for k in range(max_off + 1)], fontproperties=_font(11))
    ax3.set_yticks(range(len(cohorts)))
    ax3.set_yticklabels([f"{c['cohort_month']} (n={c['customers']:,})" for c in cohorts],
                        fontproperties=_font(10))
    for r, c in enumerate(cohorts):
        for k, v in enumerate(c["retention"]):
            ax3.text(k, r, f"{v:.0%}", ha="center", va="center",
                     fontproperties=_font(9), color="#20242E" if v < 0.6 else "white")
    ax3.set_title("Cohort retention — repeat-purchase by signup month "
                  "(M0 = 100% by construction)",
                  fontproperties=_font(15, "bold"), color=INK, loc="left", pad=10)
    fig.colorbar(im, ax=ax3, fraction=0.025, pad=0.01)

    fig.text(0.06, 0.02,
             "Real public data — no synthetic values in this view. "
             "Provenance: data/REAL_DATA_PROVENANCE.md. Observational, not causal.",
             fontproperties=_font(11), color=MUTE)

    fig.savefig(PREVIEW_PATH, facecolor="white")
    plt.close(fig)


def build_workbook(result: dict[str, object]) -> None:
    WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    head = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F8A70")

    def write_sheet(title: str, headers: list[str], rows: list[list[object]]) -> None:
        ws = wb.create_sheet(title)
        for j, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=j, value=neutralize_spreadsheet_formula(str(h)))
            c.font = head
            c.fill = fill
        for i, row in enumerate(rows, 2):
            for j, val in enumerate(row, 1):
                v = neutralize_spreadsheet_formula(val) if isinstance(val, str) else val
                ws.cell(row=i, column=j, value=v)

    wb.remove(wb.active)
    write_sheet(
        "RFM segments",
        ["Segment", "Customers", "Share", "Revenue EUR", "Automation flow", "Trigger"],
        [[s["segment"], s["customers"], s["customer_share"], s["revenue_eur"],
          s["automation_flow"], s["trigger"]] for s in result["rfm"]["segments"]],
    )
    write_sheet(
        "CLV by country",
        ["Country", "Customers", "Orders/customer", "AOV EUR", "Historical CLV EUR"],
        [[c["country"], c["customers"], c["orders_per_customer"],
          c["avg_order_value_eur"], c["historical_clv_eur"]]
         for c in result["clv_by_country"]["ranked"]],
    )
    co = result["cohort_retention"]
    write_sheet(
        "Cohort retention",
        ["Cohort", "Customers"] + [f"M{k}" for k in range(co["max_offset"] + 1)],
        [[c["cohort_month"], c["customers"], *c["retention"]] for c in co["cohorts"]],
    )
    wb.save(WORKBOOK_PATH)


def main() -> None:
    result = crm_retention.run()
    build_preview(result)
    build_workbook(result)
    print(f"Wrote {PREVIEW_PATH.relative_to(ROOT)} and {WORKBOOK_PATH.relative_to(ROOT)} "
          f"from REAL Online Retail II data.")


if __name__ == "__main__":
    main()
