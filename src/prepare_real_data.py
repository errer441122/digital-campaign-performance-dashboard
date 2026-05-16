"""Prepare a committable orders sample from the **real** public dataset
*Online Retail II* (UCI ML Repository, ID 502).

Source : https://archive.ics.uci.edu/static/public/502/online+retail+ii.zip
License: Creative Commons Attribution 4.0 International (CC BY 4.0).
         Chen, D. (2019). Online Retail II. UCI Machine Learning Repository.

This is real UK online-retail transaction data (Dec 2009 - Dec 2011). The
script downloads the source once (SHA256-verified), cleans it with stated
rules, aggregates to order level, and writes a small, deterministic CSV that
the CRM analyses consume. Only the prepared CSV + provenance are committed;
the 45 MB source is cached locally and git-ignored.

Cleaning rules (documented, conservative for RFM/CLV):
  - keep rows with a Customer ID;
  - keep Quantity > 0 and Price > 0 (drops returns/credits and zero lines);
  - drop cancellation invoices (Invoice starting with 'C');
  - aggregate to one row per (Customer ID, Invoice).
"""

from __future__ import annotations

import csv
import hashlib
import urllib.request
import zipfile
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
CACHE_DIR = ROOT / ".cache"
SOURCE_XLSX = CACHE_DIR / "online_retail_II.xlsx"
SOURCE_ZIP = CACHE_DIR / "online_retail_II.zip"
OUT_CSV = ROOT / "data" / "online_retail_orders.csv"
PROVENANCE = ROOT / "data" / "REAL_DATA_PROVENANCE.md"

SRC_URL = "https://archive.ics.uci.edu/static/public/502/online+retail+ii.zip"
SRC_ZIP_SHA256 = "572e36277c2390fbfde10664750731e0a86f55e33470d91919085f0408e67bfb"


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def ensure_source() -> None:
    if SOURCE_XLSX.exists():
        return
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    if not SOURCE_ZIP.exists():
        print(f"Downloading {SRC_URL} ...")
        urllib.request.urlretrieve(SRC_URL, SOURCE_ZIP)  # noqa: S310 (pinned UCI URL)
    digest = _sha256(SOURCE_ZIP)
    if digest != SRC_ZIP_SHA256:
        raise SystemExit(
            f"Source checksum mismatch: expected {SRC_ZIP_SHA256}, got {digest}"
        )
    with zipfile.ZipFile(SOURCE_ZIP) as zf:
        zf.extractall(CACHE_DIR)
    if not SOURCE_XLSX.exists():
        raise SystemExit("Expected online_retail_II.xlsx inside the source zip")


def _norm_customer(value: object) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, float):
        return str(int(value))
    return str(value).strip()


def build_orders() -> tuple[list[dict[str, object]], int, int]:
    wb = openpyxl.load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    # order key -> aggregated record
    orders: dict[tuple[str, str], dict[str, object]] = {}
    raw_rows = 0
    kept_rows = 0

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        next(rows)  # header
        for invoice, _stock, _desc, qty, inv_date, price, cust, country in rows:
            raw_rows += 1
            customer = _norm_customer(cust)
            if not customer:
                continue
            inv = str(invoice).strip()
            if inv.startswith("C"):  # cancellation
                continue
            try:
                quantity = int(qty)
                unit_price = float(price)
            except (TypeError, ValueError):
                continue
            if quantity <= 0 or unit_price <= 0 or inv_date is None:
                continue
            kept_rows += 1
            key = (customer, inv)
            rec = orders.get(key)
            line_value = quantity * unit_price
            if rec is None:
                orders[key] = {
                    "customer_id": customer,
                    "order_id": inv,
                    "order_date": inv_date.date(),
                    "country": str(country).strip() if country else "Unspecified",
                    "order_value_eur": line_value,
                    "n_items": quantity,
                }
            else:
                rec["order_value_eur"] += line_value
                rec["n_items"] += quantity
                if inv_date.date() < rec["order_date"]:
                    rec["order_date"] = inv_date.date()

    # Per-customer first purchase -> signup / cohort.
    first_date: dict[str, object] = {}
    for rec in orders.values():
        c = rec["customer_id"]
        d = rec["order_date"]
        if c not in first_date or d < first_date[c]:
            first_date[c] = d

    out: list[dict[str, object]] = []
    for rec in orders.values():
        signup = first_date[rec["customer_id"]]
        out.append(
            {
                "customer_id": rec["customer_id"],
                "order_id": rec["order_id"],
                "order_date": rec["order_date"].isoformat(),
                "cohort_month": signup.isoformat()[:7],
                "signup_date": signup.isoformat(),
                "country": rec["country"],
                "order_value_eur": round(float(rec["order_value_eur"]), 2),
                "n_items": int(rec["n_items"]),
            }
        )
    out.sort(key=lambda r: (r["customer_id"], r["order_date"], r["order_id"]))
    return out, raw_rows, kept_rows


def write_provenance(rows: list[dict[str, object]], raw: int, kept: int) -> None:
    customers = sorted({r["customer_id"] for r in rows})
    dates = [r["order_date"] for r in rows]
    countries = len({r["country"] for r in rows})
    PROVENANCE.write_text(
        f"""# Real Data Provenance

## Source

- **Dataset:** Online Retail II — UCI Machine Learning Repository (ID 502)
- **URL:** {SRC_URL}
- **Citation:** Chen, D. (2019). *Online Retail II*. UCI Machine Learning
  Repository. https://doi.org/10.24432/C5CG6D
- **License:** Creative Commons Attribution 4.0 International (CC BY 4.0).
  Used here unmodified-in-spirit for a non-commercial analytics portfolio,
  with attribution.
- **Source archive SHA256:** `{SRC_ZIP_SHA256}`
- **Status:** REAL public transactional data (UK online retailer,
  Dec 2009 - Dec 2011). Not simulated.

## Cleaning rules applied

1. Keep only rows with a Customer ID.
2. Keep Quantity > 0 and Price > 0 (drops returns, credits, zero lines).
3. Drop cancellation invoices (Invoice starting with `C`).
4. Aggregate to one row per (Customer ID, Invoice) = one order.

## Prepared sample

- **File:** `data/online_retail_orders.csv`
- **Raw source rows scanned:** {raw:,}
- **Rows kept after cleaning:** {kept:,}
- **Orders (rows in prepared file):** {len(rows):,}
- **Distinct customers:** {len(customers):,}
- **Distinct countries:** {countries}
- **Order date range:** {min(dates)} … {max(dates)}
- **Prepared file SHA256:** `{_sha256(OUT_CSV)}`

Re-running `python src/prepare_real_data.py` reproduces this file byte-for-byte
from the SHA256-pinned source.

## Hybrid disclosure

The CRM analyses (RFM, cohort retention, CLV, customer lifecycle) run on
**this real dataset**. The campaign-performance, multi-touch attribution,
budget-reallocation, A/B and lead-scoring-engagement / GDPR-consent modules
have **no equivalent in any permissive public dataset** and remain
**clearly-labelled simulated** data — see `data/DATA_CARD.md`.
""",
        encoding="utf-8",
    )


def main() -> None:
    ensure_source()
    rows, raw, kept = build_orders()
    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    # Explicit LF so the file is byte-identical locally, in git (eol=lf) and
    # in CI — the provenance SHA256 gate depends on it.
    with OUT_CSV.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(rows[0].keys()), lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
    write_provenance(rows, raw, kept)
    print(
        f"Wrote {len(rows):,} orders for "
        f"{len({r['customer_id'] for r in rows}):,} customers "
        f"to {OUT_CSV.relative_to(ROOT)} (real Online Retail II)."
    )


if __name__ == "__main__":
    main()
